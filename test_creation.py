from random import choice
from string import ascii_letters

import openpyxl

data_dir = './data'
text_length = 5000
row_count = 50

wb = openpyxl.load_workbook(f'{data_dir}/form.xlsx')
wba = wb.active

for row in range(row_count):
    texts = [''.join(choice(ascii_letters) for i in range(text_length)) for j in range(4)]
    wba.cell(row=row+2, column=1).value = row+1
    for j in range(2, 6):
        wba.cell(row=row+2, column=j).value = texts[j - 2]

wb.save('./data/test.xlsx')
